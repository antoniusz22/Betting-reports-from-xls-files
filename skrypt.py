from openpyxl import load_workbook


def report(file):

    num_of_match = int(input("Z ktorej kolejki chcesz stworzyc raport: "))
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active
    for points in sheet.iter_rows(min_row=46, max_row=46,  min_col=4, max_col=60, values_only=True):
        for names in sheet.iter_rows(min_row=47, max_row=47, min_col=4, max_col=60, values_only=True):
            print(points)
            print(names)
    for kolejka in sheet.iter_rows(min_row=2 + num_of_match, max_row=2 + num_of_match, min_col=4, max_col=60, values_only=True):
        print(kolejka)
    with open('raport.txt', 'w') as file:
        for i in range(len(points)):
            file.write("@" + str(names[i]) + " - ")
            file.write(str(points[i]) + " ")
            file.write("(+" + str(kolejka[i]) + ")"+ "\n")
    print("Tworze raport...")
    print("Raport gotowy w pliku 'raport.txt'")


report('TYPER-19.xlsx')