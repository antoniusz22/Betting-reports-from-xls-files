from openpyxl import load_workbook


def report(file):
    file = input("Podaj nazwe pliku")
    num_of_match = int(input("Z ktorej kolejki chcesz stworzyc raport: "))
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active
    for points in sheet.iter_rows(min_row=40, max_row=40,  min_col=4, max_col=47, values_only=True):
        for names in sheet.iter_rows(min_row=41, max_row=41, min_col=4, max_col=47, values_only=True):
            print(points)
            print(names)
    for kolejka in sheet.iter_rows(min_row=2 + num_of_match, max_row=2 + num_of_match, min_col=4, max_col=47, values_only=True):
        print(kolejka)
    with open('raport.txt', 'w') as file:
        for i in range(len(points)):
            file.write("@" + str(names[i]) + " - ")
            file.write(str(points[i]) + " ")
            file.write("(+" + str(kolejka[i]) + ")"+ "\n")
    print("Tworze raport...")
    print("Raport gotowy w pliku 'raport.txt'")


report('TYPER-24.xlsm')